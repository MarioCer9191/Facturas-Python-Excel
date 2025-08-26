import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers

# Rutas
RAIZ = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
CSV_ENTRADA = os.path.join(RAIZ, "1PDF", "FACT3.csv")
CARPETA_SALIDA = os.path.join(RAIZ)
EXCEL_SALIDA = os.path.join(RAIZ, "FACTURAS.xlsx")

# ==========================
# 1. Cargar y procesar datos
# ==========================
df = pd.read_csv(CSV_ENTRADA)

# Normalizar nombre de columnas (por si vienen con espacios)
df.columns = [c.strip().upper() for c in df.columns]

# Filtrar facturas: excluir "NO DEDUCIBLES"
if "CATEGORIA" in df.columns:
    df = df[df["CATEGORIA"].str.upper() != "NO DEDUCIBLES"]

# Convertir fechas y dar formato DD-MM-YYYY
df["FECHA"] = pd.to_datetime(df["FECHA"], errors="coerce")

# Si no existe la columna CONCEPTO, usar NOMBREEMISOR
if "CONCEPTO" not in df.columns:
    df["CONCEPTO"] = df.get("NOMBREEMISOR", "")

# Ordenar por fecha y luego por CONCEPTO
df = df.sort_values(by=["FECHA", "CONCEPTO"]).reset_index(drop=True)

# ==========================
# 2. Agrupar y subtotales por CATEGORIA
# ==========================
columnas_num = ["IMPORTE", "IVA", "TOTAL"]
filas_finales = []

if "CATEGORIA" in df.columns:
    for categoria, grupo in sorted(df.groupby("CATEGORIA"), key=lambda x: x[0]):
        for row in grupo.to_dict("records"):
            filas_finales.append(row)
        subtotal = {col: grupo[col].sum() for col in columnas_num}
        subtotal.update({col: "" for col in df.columns if col not in columnas_num and col != "CATEGORIA"})
        subtotal["CONCEPTO"] = f"SUBTOTAL {categoria}"
        filas_finales.append(subtotal)
        filas_finales.append({col: "" for col in df.columns if col != "CATEGORIA"})
else:
    filas_finales.extend(df.to_dict("records"))

df_final = pd.DataFrame(filas_finales)

# Formatear columnas numéricas a 2 decimales (incluyendo ceros)
for col in columnas_num:
    if col in df_final.columns:
        df_final[col] = pd.to_numeric(df_final[col], errors="coerce").fillna(0).map(lambda x: f"{x:.2f}")

# Formato fecha DD-MM-YYYY
if "FECHA" in df_final.columns:
    df_final["FECHA"] = pd.to_datetime(df_final["FECHA"], errors="coerce").dt.strftime("%d-%m-%Y")

# Guardar en Excel sin formato
df_final.to_excel(EXCEL_SALIDA, index=False)

# ==========================
# 3. Dar formato con openpyxl
# ==========================
wb = load_workbook(EXCEL_SALIDA)
ws = wb.active

# Estilos
header_fill = PatternFill("solid", fgColor="4F81BD")  # azul
header_font = Font(bold=True, color="FFFFFF")
subtotal_fill = PatternFill("solid", fgColor="D9D9D9")
subtotal_font = Font(bold=True)
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Encabezados
for col in range(1, ws.max_column + 1):
    cell = ws.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Contenido y formato numérico
for row in range(2, ws.max_row + 1):
    concepto = str(ws.cell(row=row, column=ws.max_column - 5).value).upper() if ws.cell(row=row, column=ws.max_column - 5).value else ""
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = thin_border
        # Columnas de números (IMPORTE, IVA, TOTAL) siempre con 2 decimales
        if ws.cell(row=1, column=col).value in columnas_num:
            cell.number_format = numbers.FORMAT_NUMBER_00
            cell.alignment = Alignment(horizontal="right")
        # Subtotales
        if concepto.startswith("SUBTOTAL"):
            cell.fill = subtotal_fill
            cell.font = subtotal_font

# Ajustar ancho de columnas automáticamente
for col in ws.columns:
    max_length = 0
    column = col[0].column
    column_letter = get_column_letter(column)
    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    ws.column_dimensions[column_letter].width = max_length + 2

# Guardar
wb.save(EXCEL_SALIDA)
print("Archivo generado:", EXCEL_SALIDA)