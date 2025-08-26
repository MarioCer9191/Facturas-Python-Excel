import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers

# Rutas
RAIZ = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
EXCEL_ENTRADA = os.path.join(RAIZ, "FACTURAS.xlsx")
EXCEL_SALIDA = os.path.join(RAIZ, "FACTURAS_ORGANIZADAS.xlsx")

# ==========================
# 1. Cargar y procesar datos
# ==========================
df = pd.read_excel(EXCEL_ENTRADA)

# Eliminar las columnas CLAVE PRODUCTO y CATEGORIA si existen
for col_drop in ["CLAVE PRODUCTO", "CATEGORIA"]:
    if col_drop in df.columns:
        df = df.drop(columns=[col_drop])

# Eliminar "0.00" en columnas numéricas de filas vacías (sin datos)
columnas_num = ["IMPORTE", "IVA", "TOTAL"]
columnas_info = [col for col in df.columns if col not in columnas_num]
for col in columnas_num:
    if col in df.columns:
        df[col] = df.apply(
            lambda row: "" if (
                all(pd.isna(row[c]) or str(row[c]).strip() == "" for c in columnas_info) and
                str(row[col]).strip() == "0.00"
            ) else row[col], axis=1
        )

# Guardar en Excel sin formato
df.to_excel(EXCEL_SALIDA, index=False)

# ==========================
# 2. Dar formato con openpyxl
# ==========================
wb = load_workbook(EXCEL_SALIDA)
ws = wb.active

# Estilos
header_fill = PatternFill("solid", fgColor="4F81BD")  # azul
header_font = Font(bold=True, color="FFFFFF")
subtotal_fill = PatternFill("solid", fgColor="D9D9D9")
subtotal_font = Font(bold=True)
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Encabezados
for col in range(1, ws.max_column + 1):
    cell = ws.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Contenido y formato numérico/texto
for row in range(2, ws.max_row + 1):
    concepto = str(ws.cell(row=row, column=ws.max_column - 3).value).upper() if ws.cell(row=row, column=ws.max_column - 3).value else ""
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = thin_border
        # Columnas de números (IMPORTE, IVA, TOTAL) siempre con 2 decimales
        if ws.cell(row=1, column=col).value in columnas_num:
            if cell.value == "" or cell.value is None:
                cell.number_format = numbers.FORMAT_TEXT
            else:
                cell.number_format = numbers.FORMAT_NUMBER_00
                cell.alignment = Alignment(horizontal="right")
        else:
            cell.number_format = numbers.FORMAT_TEXT
        # Subtotales
        if concepto.startswith("SUBTOTAL"):
            cell.fill = subtotal_fill
            cell.font = subtotal_font

# Ajustar ancho de columnas automáticamente
for col in ws.columns:
    max_length = 0
    column = col[0].column
    column_letter = get_column_letter(column)
    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    ws.column_dimensions[column_letter].width = max_length + 2

# Guardar cambios
wb.save(EXCEL_SALIDA)
print(f"✅ Archivo Excel generado y formateado: {EXCEL_SALIDA}")